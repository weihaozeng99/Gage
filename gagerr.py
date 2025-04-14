import streamlit as st
import pandas as pd
import re, io
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import traceback

def get_row_num(cell):
    return int(re.search(r'\d+', cell).group())

def get_column_num(cell):
    return column_index_from_string(re.search(r'[A-Za-z]+', cell).group())

def point_distance(point1, point2):
    return get_row_num(point2) - get_row_num(point1) + 1

#TODO: Need to confrim the data cell base
data_cell_base = 'C13'
template_data_offset = 3

st.title('GAGERR')

# Upload RFQ then create new template 
rfq = st.file_uploader("Upload RFQ")
if rfq is not None:
    try:
        df = pd.read_excel(rfq)
        st.success('RFQ uploaded successfully')
        # Input points and limits


        

    except Exception as e:
        st.error(e)
        st.stop()


    with st.form(key='rfq_form'):
        st.write('Input Operator names')
        operator_names = st.text_area('Operator Names: A, B, C')
        st.write('Input Part Number')
        part_num = st.text_area('Part Number:')
        st.write('Input Points and Limits')
        points_size = st.text_area('Size Points Location: A18-A30')

        points_form = st.text_area('Form Points Location: A18-A30')

        points_cruve = st.text_area('Cruve Points Location: A18-A30')

        limit_offset = st.text_area('Distance to tolerance, no merged cells are allowed: 4')
                    
        submit_button = st.form_submit_button(label='Submit')

    # Create new template
    if submit_button:
        st.write('Create New Template')

        # Get names
        operator_name = operator_names.split(',')


        wb = load_workbook(rfq)
        ws = wb.active
        template = load_workbook('./template.xlsx')
        # template_ws = template.active
        original_sheet = template.worksheets[1]

        size_points = points_size.split('-')
        size_num = point_distance(size_points[0], size_points[1])
        size = []
        for size_point in ws[size_points[0]:size_points[1]]:
            for cell in size_point:
                size.append(cell.value)
                
                new_sheet = template.copy_worksheet(original_sheet)
                new_sheet.title = str(cell.value)
                # Fill the part number
                new_sheet['E6'] = part_num
                # Fill the names
                new_sheet['E10'] = operator_name[0]
                new_sheet['K10'] = operator_name[1]
                new_sheet['Q10'] = operator_name[2]

                # Fill the limits
                limit_value = ws.cell(row=cell.row, column=cell.column + int(limit_offset)).value
                new_sheet['E8'] = limit_value
                clean_limit = limit_value.replace("±", "")
                new_sheet['I30'] = float(clean_limit) * 2
        # Get Form Points
        form_points = points_form.split('-')
        form_num = point_distance(form_points[0], form_points[1])
        form = []

        for form_point in ws[form_points[0]:form_points[1]]:
            for cell in form_point:
                form.append(cell.value)
                new_sheet = template.copy_worksheet(original_sheet)
                new_sheet.title = str(cell.value)
                
                new_sheet['E10'] = operator_name[0]
                new_sheet['K10'] = operator_name[1]
                new_sheet['Q10'] = operator_name[2]

                # Fill the limits
                limit_value = ws.cell(row=cell.row, column=cell.column + int(limit_offset)).value
                new_sheet['E8'] = limit_value
                clean_limit = limit_value.replace("±", "")
                new_sheet['I30'] = float(clean_limit) * 2

        # Get Cruve Points
        cruve_points = points_cruve.split('-')
        cruve_num = point_distance(cruve_points[0], cruve_points[1])
        cruve = []

        for cruve_point in ws[cruve_points[0]:cruve_points[1]]:
            for cell in cruve_point:
                cruve.append(cell.value)
                new_sheet = template.copy_worksheet(original_sheet)
                new_sheet.title = str(cell.value)
                
                new_sheet['E10'] = operator_name[0]
                new_sheet['K10'] = operator_name[1]
                new_sheet['Q10'] = operator_name[2]

                # Fill the limits
                limit_value = ws.cell(row=cell.row, column=cell.column + int(limit_offset)).value
                new_sheet['E8'] = limit_value
                clean_limit = limit_value.replace("±", "")
                new_sheet['I30'] = float(clean_limit) * 2
        

        # for i in range(size_num):
        #     # Record each point and copy a new sheet 
        #     size.append(r)



        st.write('Template created successfully')
        template.save('temp.xlsx')
        # output = io.BytesIO()
        # template.save(output)

        # output.seek(0)
        # st.download_button(label='Download Template', data=output, file_name='template.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        st.write('Upload Data')
  

# Upload Data then fill into template
data = st.file_uploader("Upload Data, .xlsx files only. Remove all ROC points")
if data is not None:
    try:
        wb_data = load_workbook(data)
        ws_data = wb_data.active
        st.success('Data uploaded successfully')
        # Fill into template
        with st.form(key='data_form'):
            st.write('Input First Points Location')
            point_loc = st.text_area('First Points Location: C2')

            data_offset = st.text_area('Distance to Absoulte Value: 3')

            points_offset = st.text_area('Distance to Next Point: 10')
                        
            submit_button = st.form_submit_button(label='Submit')
        if submit_button:
            st.write('Filling Data into Template')
            # Get First Points
            first_point = str(point_loc)
            offset = int(data_offset)
            points_offset = int(points_offset)
            # Fill data into template
            point = ws_data[first_point]
            point_value = point.value
            
            template_wb = load_workbook('temp.xlsx')
            
            # TODO: Fill up one sheet first then move to the next sheet
            # TODO: Match the point in template
            point_count = 0
            while point_value is not None:
                # st.write(point_value)
                sheet_name = str(point_value)

                st.write(sheet_name)
                template_ws = template_wb[str(point_value)]
                data_base_point = template_ws[data_cell_base]
                
                # if sheet_name in template_wb.sheetnames:
                #     template_ws = template_wb[str(point_value)]
                #     data_base_point = template_ws[data_cell_base]
                # else:
                #     point = ws_data.cell(row=2, column=point.column + points_offset)
                #     point_value = point.value
                #     st.write(point_value)
                #     point_count += 1
                       
                #     continue

                # TODO: Fill the data into the template, 9 points in a row
                point_count = 0
                for i in range(10):
                    for j in range(9):
                        # TODO: Use the offset to fill the data
                        data_cell = ws_data.cell(row=(point.row + point_count), column=point.column + int(data_offset))
                        # st.write('In data: ', data_cell.row, data_cell.column, data_cell.value)
                        if(j > 5):
                            #TODO:
                            offset_cell = data_base_point.offset(row=i, column=j+template_data_offset * 2)
                        elif (j > 2):
                            #TODO:
                            offset_cell = data_base_point.offset(row=i, column=j+template_data_offset)
                        else:
                            offset_cell = data_base_point.offset(row=i, column=j)
                        offset_cell.value = data_cell.value
                        # st.write('In template: ', offset_cell.row, offset_cell.column, offset_cell.value)
                        point_count += 1

                point = ws_data.cell(row=2, column=point.column + points_offset)
                point_value = point.value
            
            if 'sheet1' in template_wb.sheetnames:
                template_wb.remove(template_wb['sheet1'])
            
            result_sheet = template_wb.create_sheet('Result')
            result_keys = ['Acceptable', 'Margin', 'Unacceptable']
            results = {result_key: [] for result_key in result_keys}
            result_cell = 'T43'

            for ws in template_wb.worksheets:
                # Get the value of the cell in the current worksheet
                cell_value = float(ws[result_cell].value)
                # Store the value in the dictionary with the sheet name as the key
                if cell_value > 30:
                    results['Unacceptable'].append(ws.title)
                elif cell_value > 10:
                    results['Margin'].append(ws.title)
                else:
                    results['Acceptable'].append(ws.title)
            # Display the results
            headers = list(data.keys())
            for col_index, header in enumerate(headers, start=1):
                template[result_sheet].cell(row=1, column=col_index, value=header)

            # Determine the maximum number of rows needed (based on the longest list).
            max_rows = max(len(values) for values in data.values())

            # Fill in each column with the corresponding values.
            for row_index in range(1, max_rows + 1):
                for col_index, header in enumerate(headers, start=1):
                # Use row_index - 1 because lists are 0-indexed.
                    try:
                        cell_value = data[header][row_index - 1]
                    except IndexError:
                        cell_value = None  # If this list is shorter, you can leave the cell empty.
                    template[result_sheet].cell(row=row_index + 1, column=col_index, value=cell_value)

                
            st.write('Data filled successfully')
            template_wb.save('filled_template.xlsx')
            with open('filled_template.xlsx', 'rb') as f:
                download_data = f.read()
            st.download_button(label='Download Filled Template', data=download_data, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    except Exception as e:
        st.error(traceback.format_exc())
        st.stop()

